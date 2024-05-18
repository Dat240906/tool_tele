from concurrent.futures import thread
import threading
import requests
from bs4 import  BeautifulSoup
import telebot
from datetime import datetime
from time import sleep
import copy
import openpyxl, tempfile
from telebot import types
TOKEN = '6525705295:AAGaoXer8twK9ZpvUDXyWVclGA3WD6ioH-0'
bot = telebot.TeleBot(TOKEN)

TIME_XSMB = (18, 33, 0)
TIME_STOP_BET = (17, 0, 0)
TIME_SEND_EXCEL = (17, 0, 0)
USERNAME_ADMIN = 'AI_technology_production'
XSMB = {} # 5
XSMB_temp = {} # 2
users = {}
users_temp = {}
is_bet = True
current_time = datetime.now()
if 18 > int(current_time.hour) > TIME_STOP_BET[0] and 40 > int(current_time.minute) > TIME_STOP_BET[1]:
    is_bet = False


def reset_in_day():
    global users, users_temp, is_bet
    users = {}
    users_temp = {}
    is_bet = True

def handel_by_day(box_ketqua):
    time = box_ketqua.find('h2')
    time_split = time.text.split()
    if "DÒ" in time_split and "VÉ" in time_split and "XSMB" in time_split:
        return 
    title = "Xổ Số Miền Bắc"
    try:
        day_month = [time_split[i] for i in range(len(time_split)) if "/" in time_split[i]][0]
        day, month = day_month.split('/')

        if month.startswith('0'):
            month = int(month[1])
        else:
            month = int(month)
        if day.startswith('0'):
            day = int(day[1])
        else:
            day = int(day)

        day_month = f"{day}/{month}"
    except: return 

    thu = ''
    for i in range(len(time_split)):
        if "(" in time_split[i]:
            thu += f"{time_split[i][1:]} "           
        if ")" in time_split[i]:
            thu += time_split[i][:len(time_split[i])-1]
    giai = box_ketqua.find_all('tr')
    list_prize_original = []
    

       

    for item in giai:
        prize = item.find_all('td', attrs={'title': False})
        if prize:
            list_prize_original.append(prize[0].text)

    DB = list_prize_original[0]
    G1 = list_prize_original[1]

    G2_result = list_prize_original[2].split()
    G2 = [G2_result[0],G2_result[1] ]

    G3_result = list_prize_original[3].split()
    G3 = [G3_result[0], G3_result[1],G3_result[2][:5], G3_result[2][5:], G3_result[3], G3_result[4]  ]

    G4_result = list_prize_original[5].split()
    G4 = [G4_result[0], G4_result[1], G4_result[2], G4_result[3]]

    G5_result = list_prize_original[6].split()
    G5 = [G5_result[0], G5_result[1], G5_result[2][:4], G5_result[2][4:], G5_result[3], G5_result[4]]

    G6_result = list_prize_original[8].split()
    G6 = [G6_result[0], G6_result[1], G6_result[2]]

    G7_result = list_prize_original[9].split()
    G7 = [G7_result[0], G7_result[1], G7_result[2], G7_result[3]]

    XSMB[day_month] = {
        "DB":DB, 
        "G1":G1,
        "G2":G2,
        "G3":G3,
        "G4":G4,
        "G5":G5,
        "G6":G6,
        "G7":G7,
    }  

def export_excel():
    time = (datetime.now().day, datetime.now().month)
    if not users:
        return 
    wb = openpyxl.Workbook()
    sheet = wb.active

    gray_fill = openpyxl.styles.PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    white_font = openpyxl.styles.Font(color="FFFFFF")
    for cell in ["A1", "B1", "C1", "D1", "E1"]:
        sheet[cell].fill = gray_fill
        sheet[cell].font = white_font

    sheet['A1'] = 'Thời gian'
    sheet['B1'] = 'Tên khách hàng'
    sheet['C1'] = 'Số đặt'
    sheet['D1'] = 'Lô/Đề'
    sheet['E1'] = 'Tiền đặt'

    column_widths = [15, 30, 10, 10, 15]
    for i, width in enumerate(column_widths):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width

    chat_id = ''
    i=2
    for id, user_ in users.items():
        for id_user, info_user in user_.items():
            chat_id = info_user[0][3]
            time = info_user[0][5]
            name = f"{info_user[0][0]} {info_user[0][1]}"
            number_bet = info_user[1]
            lo_or_de = info_user[2]
            sheet[f'A{i}'] = time
            sheet[f'B{i}'] = name
            sheet[f'C{i}'] = number_bet
            sheet[f'D{i}'] = lo_or_de
            i+=1
    

    # Lưu file Excel tạm thời
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        wb.save(temp_file.name)

        # Gửi file Excel tạm thời qua bot
        with open(temp_file.name, 'rb') as file:
            bot.send_document(chat_id, file, caption=f'Dữ liệu đặt cược.')
def tong_duyet(day_month):
    global XSMB_temp, users_temp
    XSMB_temp = copy.deepcopy(XSMB)
    data = XSMB_temp[day_month]
    DB = data['DB'][len(data['DB'])-2:]
    XSMB_temp[day_month]['DB'] = DB
    G1 = data['G1'][len(data['G1'])-2:]
    XSMB_temp[day_month]['G1'] = G1
    G2_list = data['G2']
    for i in range(len(G2_list)):
        G2_list[i] = G2_list[i][len(G2_list[i])-2:]
    G3_list = data['G3']
    for i in range(len(G3_list)):
        G3_list[i] = G3_list[i][len(G3_list[i])-2:]
    G4_list = data['G4']
    for i in range(len(G4_list)):
        G4_list[i] = G4_list[i][len(G4_list[i])-2:]
    G5_list = data['G5']
    for i in range(len(G5_list)):
        G5_list[i] = G5_list[i][len(G5_list[i])-2:]
    G6_list = data['G6']
    for i in range(len(G6_list)):
        G6_list[i] = G6_list[i][len(G6_list[i])-2:]
    G7_list = data['G7']
    for i in range(len(G7_list)):
        G7_list[i] = G7_list[i][len(G7_list[i])-2:]
    
    users_temp = copy.deepcopy(users)
    for id, users_ in users_temp.items():
        for id_user, info in users_.items():
            number_ = info[1]
            type_ = info[2]
            if type_ == 'đề':
                if number_ == DB:
                    number_ = "Đặc Biệt"
            elif type_ == 'lô':
                if number_ == G1:
                    number_ = "Nhất"
                elif number_ in G2_list:
                    number_ = "Nhì"
                elif number_ in G3_list:
                    number_ = "Ba"
                elif number_ in G4_list:
                    number_ = "Tư"
                elif number_ in G5_list:
                    number_ = "Năm"
                elif number_ in G6_list:
                    number_ = "Sáu"
                elif number_ in G7_list:
                    number_ = "Bảy"
          
            users_temp[id][id_user][1] = number_




def get_ketqua():
    url = "https://xskt.com.vn/xsmb"


    response = requests.get(url)
    suop = BeautifulSoup(response.content, "html.parser")

    box_ketqua = suop.find_all('div', class_="box-ketqua")
    for item in box_ketqua:
        if item:
            handel_by_day(item)
    
get_ketqua()
def add_number(info, number, lo_or_de):
    first_name = info[0]
    last_name = info[1]
    id_user = info[2]
    chat_id = info[3]
    message = info[4]
    time = info[5]
    for user_ in users.values():
        data_id_user = next(iter(user_.keys()))
        data_value = next(iter(user_.values()))
        data_number = data_value[1]
        data_lo_or_de = data_value[2]
        if data_id_user == id_user and data_number == number and data_lo_or_de == lo_or_de:
            return 
        else:
            print(data_id_user)
            print(id_user)
            print(data_number)
            print(number)
            print(data_lo_or_de)
    try:
        users[f"{chat_id}_{message.message_id}"]
    except KeyError:
        users[f"{chat_id}_{message.message_id}"] = {}
        
    try:
        user = users[f"{chat_id}_{message.message_id}"][id_user]
        print(user)
        numbers_user = user[1]
        
        # for item in numbers_user:
        #     if int(number) == int(item[0]) and lo_or_de == item[1]:
        #         return
        numbers_user.append(number, lo_or_de)
        users[f"{chat_id}_{message.message_id}"][id_user][1] = numbers_user
    except KeyError:
        users[f"{chat_id}_{message.message_id}"][id_user] = [[first_name, last_name, id_user, chat_id, message, time], number, lo_or_de]



def get_info_user(message):
    f_n = message.from_user.first_name
    l_n = message.from_user.last_name
    if f_n is None:
        f_n = ''
    if l_n is None:
        l_n = ''
    id = message.from_user.id
    chat_id = message.chat.id
    return f_n, l_n, id, chat_id

def send_XSMB_result(day_month, **kwargs):
    
    data = XSMB[day_month]

    
    # bên dưới là kết quả đẹp nhưng chưa thích nghi được với điện thoại anh nhá

#     XSMB_RESULT = f"""
# <u>KẾT QUẢ XSMB {day_month}</u>
# <b>
# ╔════════════════════════════════╗
# ║ DB ║                                   {data['DB']}                               ║
# ╠════════════════════════════════╣
# ║ G1 ║                                   {data['G1']}                               ║
# ╠════════════════════════════════╣
# ║ G2 ║                            {data['G2'][0]}  {data['G2'][1]}                         ║
# ╠════════════════════════════════╣
# ║ G3 ║                   {data['G3'][0]}  {data['G3'][1]}  {data['G3'][2]}                     ║
# ║       ║                   {data['G3'][3]}  {data['G3'][4]}  {data['G3'][5]}                     ║
# ╠════════════════════════════════╣
# ║ G4 ║                 {data['G4'][0]}  {data['G4'][1]}  {data['G4'][2]}  {data['G4'][3]}                   ║
# ╠════════════════════════════════╣
# ║ G5 ║                        {data['G5'][0]}  {data['G5'][1]}  {data['G5'][2]}                       ║
# ║       ║                       {data['G5'][3]}  {data['G5'][4]}  {data['G5'][5]}                        ║
# ╠════════════════════════════════╣
# ║ G6 ║                          {data['G6'][0]}  {data['G6'][1]}  {data['G6'][2]}                           ║
# ╠════════════════════════════════╣
# ║ G7 ║                          {data['G7'][0]}  {data['G7'][1]}  {data['G7'][2]}  {data['G7'][3]}                            ║
# ╚════════════════════════════════╝
# </b>
# """

    # đã thích nghi được với điện thoại 
    XSMB_RESULT = f"""
<u>KẾT QUẢ XSMB {day_month}</u>
<b>

 DB │ {data['DB']}              
──────────────
 G1 │ {data['G1']}              
──────────────
 G2 │ {data['G2'][0]}  {data['G2'][1]}        
──────────────
 G3 │ {data['G3'][0]}  {data['G3'][1]}  {data['G3'][2]}
       │ {data['G3'][3]}  {data['G3'][4]}  {data['G3'][5]}    
──────────────
 G4 │ {data['G4'][0]}  {data['G4'][1]}  {data['G4'][2]}  {data['G4'][3]}  
──────────────
 G5 │ {data['G5'][0]}  {data['G5'][1]}  {data['G5'][2]}
       │ {data['G5'][3]}  {data['G5'][4]}  {data['G5'][5]}       
──────────────
 G6 │ {data['G6'][0]}  {data['G6'][1]}  {data['G6'][2]}          
──────────────
 G7 │ {data['G7'][0]}  {data['G7'][1]}  {data['G7'][2]}  {data['G7'][3]}           
──────────────
</b>
"""
    if "message" in kwargs:
        message = kwargs["message"]
        return bot.reply_to(message, XSMB_RESULT, parse_mode = 'html')
    list_chat_id = []
    if not users:
        return 
    for id, users_ in users.items():
        for id_user, info_user in users_.items():
            chat_id_ = info_user[0][3]
            if chat_id_ in list_chat_id:
                continue
            list_chat_id.append(info_user[0][3])
            continue

    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton('Thuê BOT tại đây!', url='https://t.me/CSKH_VPN4G_BOT'))

    for chat_id in list_chat_id:
        bot.send_message(chat_id,XSMB_RESULT, parse_mode = 'html', reply_markup=keyboard)
        sleep(1)


def send_result_game(day_month):
    global users, users_temp
    data = XSMB_temp[day_month]
    DB = data['DB'] #one
    G1 = data['G1'] #one
    G2 = data['G2']
    G3 = data['G3']
    G4 = data['G4']
    G5 = data['G5']
    G6 = data['G6']
    G7 = data['G7']

    for id, users_ in users.items():
        for id_user, info_user in users_.items():
            data_numbers_original = users[id][id_user][1]
            data_numbers_changed = users_temp[id][id_user][1]
            list_number = []
            if data_numbers_original != data_numbers_changed:
                list_number.append(data_numbers_original)
                sleep(1)
                bot.reply_to(info_user[0][4], f"<b>Chúc mừng trúng giải {data_numbers_changed.upper()}!</b>", parse_mode='html')
            # try:
            #     room = users_win_game[chat_id]
            #     room[id_user] = [info_user, list_number]
            #     users_win_game[chat_id] = room
            # except:
            #     users_win_game[chat_id] = {}
            #     users_win_game[chat_id][id_user] = [info_user, list_number]
    users = {}
    users_temp = {}
    reset_in_day()
def auto_check():
    global users_temp, is_bet
    while True:
        sleep(1)
        day_month = f"{datetime.now().day}/{datetime.now().month}"
        time = (datetime.now().hour, datetime.now().minute, datetime.now().second)
        
        if time[0] == TIME_STOP_BET[0] and time[1] == TIME_STOP_BET[1] and time[2] == TIME_STOP_BET[2]:
            is_bet = False
        if time[0] == TIME_XSMB[0] and time[1] == TIME_XSMB[1] and time[2] == TIME_XSMB[2] :
            get_ketqua()
            try:
                XSMB[day_month]
            except: continue
            tong_duyet(day_month)
            _thread_send_result_XSMB = threading.Thread(target=send_XSMB_result, args=(day_month,))
            _thread_send_result_XSMB.setDaemon(True)
            _thread_send_result_XSMB.start()

            _thread_send_result_game = threading.Thread(target=send_result_game, args=(day_month,))
            _thread_send_result_game.setDaemon(True)
            _thread_send_result_game.start()
            
            is_bet = True
        if time[0] == TIME_SEND_EXCEL[0] and time[1] == TIME_SEND_EXCEL[1] and time[2] == TIME_SEND_EXCEL[2]:
            export_excel()



@bot.message_handler(func=lambda message: True)
def echo_all(message):
    day_month = f"{datetime.now().day}/{datetime.now().month}"
    time = f'{datetime.now().hour}:{datetime.now().minute}:{datetime.now().second}'
    # day_month = "29/4"
    content = message.text.lower()
    username = message.from_user.username
    first_name, last_name, id_user, chat_id = get_info_user(message)
    if content.startswith("/lô") and is_bet:
        len_ = len(content.split())
        if len_ == 2:
            number = content.split()[1]
            add_number((first_name, last_name, id_user, chat_id, message, time), number, "lô")
            print(users)
            return bot.reply_to(message, "<b>Đặt số thành công!</b>", parse_mode = 'html')
        
        return 
    if content.startswith("/đề") and is_bet:
        len_ = len(content.split())
        if len_ == 2:
            number = content.split()[1]
            add_number((first_name, last_name, id_user, chat_id, message, time), number, "đề")
            print(users)
            return bot.reply_to(message, "<b>Đặt số thành công!</b>", parse_mode = 'html')
        return 
    if (content.startswith("/lô") or content.startswith("/đề")) and not is_bet:
        return bot.reply_to(message, "<b>Đã hết thời gian đặt số!</b>", parse_mode = 'html')

    # if "kết quả" in content and username == USERNAME_ADMIN:
    if content == 'kết quả hôm nay' and username == USERNAME_ADMIN:
        day_month = f"{datetime.now().day}/{datetime.now().month}"
        try:
            XSMB[day_month]
        except:
            return bot.reply_to(message, "<b>Chưa có kết quả hôm nay!</b>", parse_mode = 'html')
        return send_XSMB_result(day_month, message = message)
    if "kết quả" in content and len(content.split()) == 3 and username == USERNAME_ADMIN:
        day_month = content.split()[2]
        try:
            XSMB[day_month]
        except:
            return bot.reply_to(message, "<b>Không tìm thấy!</b>", parse_mode = 'html')
        return send_XSMB_result(day_month, message = message)




auto = threading.Thread(target=auto_check)
auto.daemon = True  
auto.start()


bot.infinity_polling(timeout=10, long_polling_timeout = 5)