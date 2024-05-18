import telebot
from datetime import datetime
import threading
from time import sleep
import openpyxl
import tempfile
from telebot import types
# TOKEN = '6918724484:AAEBy0h0r5zh6e3kNkTqFGG2tG2NJiCzU1Y'

#token test
TOKEN = '6525705295:AAGaoXer8twK9ZpvUDXyWVclGA3WD6ioH-0'






bot = telebot.TeleBot(TOKEN)
# group_send_noti = []

group_send_noti = [-1002096132981]
user_fine = {}
notification = """
BOT GHI PHẠT TỰ ĐỘNG TÍNH PHẠT KHI SAI PHẠM:

<b>➤ Đi wc không được quá 15 phút ( trên 15 phút không báo cáo 1 trả lời tin nhắn bot tự ghi phạt )

➤ Đi cy không được quá 10 phút ( trên 10 phút không báo cáo 1 trả lời tin nhắn bot tự ghi phạt )

➤ Không đi quá 2 người, cố ý để bot nhắc nhở thì ấn 1 trả lời tin nhắn bị nhắc của bạn chứ ko phải bot để xác nhận, 

➤ Không được xóa bất cứ tin nhắn nào, xóa tin nhắn bot vẫn Gim bạn vào danh sách theo dõi đến hẹn nó lại hiện lên nói : bạn bị phạt 20$

➤ Sau khi đi ra ngoài về trả lời tin nhắn của bạn là "1" để báo cáo, không trả lời mà gõ 1 bừa lưa, bot vẫn phạt

➤ Bot sẽ hiện lên và thông báo quy định hằng ngày vào lúc 10:00 AM , trước khi đi làm và tổng kết phạt lúc 10:30 PM hằng ngày , xuất file phạt

➤ Nếu bạn lách luật bằng cách ấn WC, Wc . wC ( không có đâu gấu ạ, bot ghi nhận hết )

➤ Hệ thống sẽ luôn tự động chạy, kể cả bạn mất mạng với mất điện thì bot cũng không hề ảnh hưởng gì nên đừng nghĩ việc rút mạng</b>"""

users = {}
# ví dụ:
# {id:('wc', time), [...], [...]}

TIME_SEND_NOTI = (10, 0, 0) 
TIME_SEND_EXCEL = (22, 30, 0) 

# thời gian tối đa 1 người được đi vệ sinh/hút thuốc chẳng hạn (tính theo số phút)
#ra ngoài
# MAX_TIME_WC = 15
MAX_TIME_WC = 0.2
#hút thuốc
# MAX_TIME_CY = 10
MAX_TIME_CY = 0.2


#tiền phạt (tính theo $)
FINE = 10

# số người tối đa được đi hút/vs (tính theo người)
MAX_TURN = 4



def get_info_user(message):
    f_n = message.from_user.first_name
    l_n = message.from_user.last_name
    if f_n is None:
        f_n = ''
    if l_n is None:
        l_n = ''
    id = message.from_user.id
    chat_id = message.chat.id
    return f_n, l_n, id, chat_id


def check_user(info):
    global users
    for chat_id, users_ in users.items():
        for id_user, info_user in users_.items():

            if info[2] == id_user and chat_id == info[3]:
                return True
    return False
        


def handle_WC(info, time, message, status):
    is_check_in_user = check_user(info)
    if is_check_in_user:
        pass
        # bot.send_message(info[3], f"<b>{info[0]} {info[1]} đang thực hiện {(data[0][3]).upper()}...</b>", parse_mode = 'html')
    else:
        try:
            users[info[3]]
        except:
            users[info[3]] = {}
        # data_one_group = {}
        # data_one_group[info[2]] = 
        users[info[3]][info[2]] = [(info[0], info[1], info[3], 'wc'), time, message.message_id, message, status]
        # msg = bot.reply_to(message, f'<b>Đồng ý, hành động WC giới hạn là {MAX_TIME_WC} phút.</b>', parse_mode = 'html')

def handle_CY(info, time, message, status):
    is_check_in_user = check_user(info)
    if is_check_in_user:
        pass
        # bot.send_message(info[3], f"<b>{info[0]} {info[1]} đang thực hiện {(data[0][3]).upper()}...</b>", parse_mode = 'html')
    else:
        try:
            users[info[3]]
        except:
            users[info[3]] = {}
        users[info[3]][info[2]] = [(info[0], info[1], info[3], 'cy'), time, message.message_id, message, status]
        # msg = bot.reply_to(message, f'<b>Đồng ý, hành động WC hạn là {MAX_TIME_WC} phút.</b>', parse_mode = 'html')

def send_message_punish_to_user(chat_id, id_user, status):
    data = users[chat_id][id_user]
    name = f'{data[0][0]} {data[0][1]}'
    if status == 'ok':
        if data[0][3].lower() == 'wc':
            if data[1] <= MAX_TIME_WC*60 and status == 'ok':
                return 
            time_fine = (datetime.now().hour, datetime.now().minute, datetime.now().second)
            try:
                bot.reply_to(data[3], f"<b>Cảm ơn {data[0][0]} {data[0][1]} góp {FINE}$ vào quỹ công ty, vì đi vệ sinh quá thời gian!</b>\n @ococ0707", parse_mode='html')
            except telebot.apihelper.ApiTelegramException as e:
                if e.result_json['description'] == 'Bad Request: message to reply not found':
                    try:
                        data_fine = user_fine[chat_id]
                        user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', "20", f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
                    except KeyError:
                        try:
                            user_fine[chat_id]
                        except:
                            user_fine[chat_id] = []
                        user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', "20", f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
                    users[chat_id].pop(id_user)
                    return bot.send_message(chat_id, f"<b>{name} đã xóa tin nhắn (sai quy định không được phép xóa tin nhắn, phạt 20$)</b>\n @ococ0707", parse_mode='html')
            try:
                data_fine = user_fine[chat_id]
                user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', FINE, f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
            except KeyError:
                try:
                    user_fine[chat_id]
                except:
                    user_fine[chat_id] = []
                user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', FINE, f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
            users[chat_id].pop(id_user)
        if data[0][3].lower() == 'cy':
            if data[1] <= MAX_TIME_CY*60:
                return 
            time_fine = (datetime.now().hour, datetime.now().minute, datetime.now().second)
            try:
                bot.reply_to(data[3], f"<b>Cảm ơn {data[0][0]} {data[0][1]} góp {FINE}$ vào quỹ công ty, vì đi hút thuốc quá thời gian!</b>\n @ococ0707", parse_mode='html')
            except telebot.apihelper.ApiTelegramException as e:
                if e.result_json['description'] == 'Bad Request: message to reply not found':
                    try:
                        data_fine = user_fine[chat_id]
                        user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', "20", f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
                    except KeyError:
                        try:
                            user_fine[chat_id]
                        except:
                            user_fine[chat_id] = []
                        user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', "20", f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
                    users[chat_id].pop(id_user)
                    return bot.send_message(chat_id, f"<b>{name} đã xóa tin nhắn (sai quy định không được phép xóa tin nhắn, phạt 20$)</b>\n @ococ0707", parse_mode='html')
            try:
                data_fine = user_fine[chat_id]
                user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', FINE, f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
            except KeyError:
                try:
                    user_fine[chat_id]
                except:
                    user_fine[chat_id] = []
                user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', FINE, f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
            users[chat_id].pop(id_user)
    elif status == 'er':
        if data[1] <= 60:
                return 
        bot.reply_to(data[3], f"<b>Cảm ơn {data[0][0]} {data[0][1]} góp {FINE}$ vào quỹ công ty, vì không báo cáo lại sau khi được nhắc!</b>", parse_mode='html')
        try:
            data_fine = user_fine[chat_id]
            user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', FINE, f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
        except KeyError:
            try:
                user_fine[chat_id]
            except:
                user_fine[chat_id] = []
            user_fine[chat_id].append([f'{data[0][0]} {data[0][1]}', FINE, f"{time_fine[0]}:{time_fine[1]}:{time_fine[2]}"])
        users[chat_id].pop(id_user)



def export_excel(id_group):
    time = (datetime.now().day, datetime.now().month)
    try: 
        user_fine[id_group]
    except:
        return bot.send_message(id_group, f'<b>Không có dữ liệu vi phạm {time[0]}/{time[1]}</b>', parse_mode='html')
    wb = openpyxl.Workbook()
    sheet = wb.active

    gray_fill = openpyxl.styles.PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    white_font = openpyxl.styles.Font(color="FFFFFF")
    for cell in ["A1", "B1", "C1"]:
        sheet[cell].fill = gray_fill
        sheet[cell].font = white_font

    sheet['A1'] = 'Tên'
    sheet['B1'] = 'Phạt'
    sheet['C1'] = 'Thời gian'

    column_widths = [30, 15, 15]
    for i, width in enumerate(column_widths):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width

    data = user_fine[id_group]
    i=2
    for user_fined in data:
        name = user_fined[0]
        fine = user_fined[1]
        time_fine = user_fined[2]
        sheet[f'A{i}'] = name
        sheet[f'B{i}'] = f'{fine} $'
        sheet[f'C{i}'] = time_fine
        i+=1
    

    # Lưu file Excel tạm thời
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        wb.save(temp_file.name)

        # Gửi file Excel tạm thời qua bot
        with open(temp_file.name, 'rb') as file:
            bot.send_document(id_group, file, caption=f'Dữ liệu vi phạm {time[0]}/{time[1]}')


def send_excel():
    global user_fine

    if not group_send_noti:
        return 
    for id_group in group_send_noti:
        export_excel(id_group)
        sleep(1)
        #reset
    user_fine = {}
def send_noti():
    global user_fine
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton('Liên hệ Admin', url='https://t.me/AI_technology_production'))

    if not group_send_noti:
        return 
    for id_group in group_send_noti:
        bot.send_message(id_group, notification, parse_mode='html', reply_markup=keyboard)
        #reset
        sleep(1)
    user_fine = {}
def auto_check_time():
    while True:
        sleep(1)
        time_send = (datetime.now().hour, datetime.now().minute, datetime.now().second)

        if time_send[0] == TIME_SEND_NOTI[0] and time_send[1] == TIME_SEND_NOTI[1] and time_send[2] == TIME_SEND_NOTI[2]:
            other_thread = threading.Thread(target=send_noti)
            other_thread.daemon = True  
            other_thread.start()
            
            
        if time_send[0] == TIME_SEND_EXCEL[0] and time_send[1] == TIME_SEND_EXCEL[1] and time_send[2] == TIME_SEND_EXCEL[2]:
            other_thread = threading.Thread(target=send_excel)
            other_thread.daemon = True  
            other_thread.start()
        if not users:
            continue
        dict_user_copy = users.copy()
        changes = {} 
        
        for chat_id, users_ in dict_user_copy.items():
            if not users_:
                continue
            for id_user, info_user in users_.items():
                time = users_[id_user][1] + 1
                changes.setdefault(chat_id, {})[id_user] = time

        for chat_id, users_ in changes.items():
            for id_user, time in users_.items():
                users[chat_id][id_user][1] = time

        for chat_id, users_ in changes.items():
            for id_user in users_:
                send_message_punish_to_user(chat_id, id_user, users[chat_id][id_user][4])

           
            
                   

# Hàm xử lý tin nhắn reply
def handle_reply_message(message):
    first_name, last_name, id_user, chat_id = get_info_user(message)
    replied_message = message.reply_to_message
    try:
        if replied_message is not None:
            if users[chat_id][id_user][2] !=  replied_message.id:
                return 
            if message.text != "1":
                return 
            check = check_user((first_name, last_name, id_user, chat_id))
            if check:
                try:
                    users[chat_id].pop(id_user)
                except:pass
                return 
    except KeyError:
        pass
# Trình xử lý tin nhắn khi có reply
@bot.message_handler(func=lambda message: message.reply_to_message is not None)
def reply_message_handler(message):
    handle_reply_message(message)


@bot.message_handler(func=lambda message: True)
def recived_message(message):
    first_name, last_name, id_user, chat_id = get_info_user(message)
    time = 0
    content = message.text
    username = message.from_user.username
    # đi vệ sinh 
    if content.lower() == "wc":
       
        try:
            num = len(users[chat_id])
        except KeyError:
            num = 0
        if num > MAX_TURN:
            check = check_user((first_name, last_name, id_user, chat_id))
            if check:
                return
            bot.reply_to(message, f'<b>Số người đã quá {MAX_TURN}, Không báo cáo lại sau 1p phạt {FINE}$</b>', parse_mode = 'html')
            return handle_WC((first_name, last_name, id_user, chat_id), time, message, 'er')
        handle_WC((first_name, last_name, id_user, chat_id), time, message, 'ok') 
        


    #hút thuốc
    if content.lower() == "cy":
       
        try:
            num = len(users[chat_id])
        except KeyError:
            num = 0
        if num > MAX_TURN:
            check = check_user((first_name, last_name, id_user, chat_id))
            if check:
                return
            bot.reply_to(message, f'<b>Số người đã quá {MAX_TURN}, Không báo cáo lại sau 1p phạt {FINE}$</b>', parse_mode = 'html')
            return handle_CY((first_name, last_name, id_user, chat_id), time, message, 'er')
        handle_CY((first_name, last_name, id_user, chat_id), time, message, 'ok')

    if content.lower() == 'khởi động':
        global group_send_noti
        if not (chat_id in group_send_noti):
            group_send_noti.append(chat_id)
        bot.reply_to(message, "<b>Khởi động: đã bật thông báo nội quy (10h sáng) và thông báo thống kê phạt (22h30 đêm) - hàng ngày</b>", parse_mode ='html')
    
    if content.lower() == 'đã ổn chưa' and username == 'AI_technology_production':
        bot.reply_to(message, "<b>Dạ ổn rồi ạ, đang hoạt động bình thường !</b>", parse_mode ='html')

print(f""" 
BOT setup:
    1, Thời gian WC: {MAX_TIME_WC} phút
    2, Thời gian CY: {MAX_TIME_CY} phút
    3, Số người tối đa: {MAX_TURN}
    4, Thời gian gửi thông báo: {TIME_SEND_NOTI[0]}:{TIME_SEND_NOTI[1]}:{TIME_SEND_NOTI[2]}
    5, Thời gian gửi excel: {TIME_SEND_EXCEL[0]}:{TIME_SEND_EXCEL[1]}:{TIME_SEND_EXCEL[2]}
    6, Tiền phạt: {FINE} $

BOT đang chạy...
"""
)
# running đa luồng với chạy bot
other_thread = threading.Thread(target=auto_check_time)
other_thread.daemon = True  
other_thread.start()

bot.infinity_polling(timeout=10, long_polling_timeout = 5)