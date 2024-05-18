from asyncio import transports
from pandas import to_timedelta
import telebot
from datetime import datetime
import openpyxl
import tempfile
from telebot import types
# TOKEN
# bot = telebot.TeleBot('7078921242:AAEJW1cKBH1AvX1rXYEN5X9RdC7NxFBDMpA')
bot = telebot.TeleBot('6525705295:AAGaoXer8twK9ZpvUDXyWVclGA3WD6ioH-0')

vietnam_flag = "🇻🇳"
china_flag = "🇨🇳"
uk_flag = "🇬🇧"

blue_color_start = "<b>"
blue_color_end = '</b>'
profit = 1.5
refund_money = 0
transactions = {}

total_revenue_ = 0
lang_vi = {
    "flag":f'{vietnam_flag}',
    'money': 'VND',
    'hello': f"Xin chào {vietnam_flag}! Tôi là robot do B02 (尹志平) làm ra để lưu trữ giao dịch nạp - rút trong ngày.\nCảm ơn bạn đã sử dụng. Rất vui được phục vụ!",
    'trade_fail': "Lỗi cú pháp!",
    'trade_success_pl': 'Giao dịch nạp tiền thành công.',
    'trade_success_sub': 'Giao dịch rút tiền thành công.',
    'data_analysis': {
        'total_bill': f"Tổng cộng số bill là",
        'profit':'Lãi suất kí gửi',
        'total_rate_now': 'Tổng tỉ giá hiện tại là',
        'total_deposit': 'Tổng tiền nạp',
        'total_withdraw': 'Tổng tiền rút',
        'total_revenue': 'Tổng khoản thu',
        'unrefund_money': 'Số tiền chưa hoàn trả',
        'refund_money': 'Số tiền đã hoàn trả',
        'total_revenue_USD': 'Tổng doanh thu tính theo USD'
    },
    'trade_fail2': "Đã có lỗi xảy ra trong quá trình xử lý giao dịch.",
    'excel_config': {
        'time': "Thời gian",
        'gd':"giao dịch",
        'user':'Người thực hiện',
        'money': 'Số tiền',
        'money_withdraw': 'Số tiền rút',
        'total_money': 'Tổng số tiền',
        'total_withdraw': 'Tổng rút',
        'value': 'Giá trị',
        'total_revenue': 'Tổng doanh thu',
        'today_analysis': 'Thống kê ngày hôm nay',
        'fail_excel':"Đã có lỗi xảy ra trong quá trình xuất file Excel.",

    },
    'refunds_message_success':"Hoàn trả thành công",
    'refunds_message_fail':"Hoàn trả thất bại, số tiền hoàn trả phải bé hơn hoặc bằng số tiền chưa hoàn trả.",
    'success_change':"Đã thay đổi thành công!",
    'change_rate':'Đã thay đổi tỉ giá thành',
    'value_er':'Nhập tỷ giá cũng sai :))) NGU!',
    'update':'Đã nâng cấp thành công. Hệ thống được cập nhật lại mọi thứ,\n Tỷ giá về mặc định là 25.000VND \n Để biết thêm về bot này vui lòng liên hệ @Zhiping_Yin \n Xin cảm ơn.',
    'help':'Cách sử dụng Robot này:\n\t/start: để bắt đầu tính bill lại từ đầu ngày\n\t/end: để kết thúc một ngày\n\t/file: để xuất file báo cáo tổng tiền nạp trong ngày\n\n\t/rate {A}: $ 1 = A VND \n\t+{số}: để thêm bill nạp tiền\n\t-{số}: để thêm bill rút tiền\n\t/refunds A VND: hoàn tiền A VND\n\n\t/help: để xem hướng dẫn\n\t/update: để cập nhật hệ thống\n\t/set {ngôn ngữ}: (ngôn ngữ hỗ trợ: tiếng việt: vi, tiếng anh: en, tiếng trung: zh)\n\t/dir A: thay đổi lãi suất kí gửi thành A%',
    'end':'Bot đã dừng lại. Hãy sử dụng lệnh /start để khởi động lại.',
    'syntax_er':'Cú pháp sử dụng không tồn tại, dùng /help để biết thêm.',
    'erorr_undefined':'Lỗi không xác định',
    'change_language': 'Thay đổi ngôn ngữ thành công.',
    'send_money':'Tiền Gửi',
    'withdraw_money':'Tiền rút',
    'refunds_money':'Tiền hoàn trả',
    'gd':"giao dịch",
    'error_float':'Sử dụng dấu chấm để ngăn cách số thập phân.'
}
lang_en = {
    "flag":f'{uk_flag}',
    'money': 'VND',
    'hello': f"Hello {uk_flag}! I'm a robot created by B02 (尹志平) to record deposit - withdrawal transactions throughout the day.\nThank you for using. Happy to serve you!",
    'trade_fail': "Syntax error",
    'trade_success_pl': ' Deposit transaction successful.',
    'trade_success_sub': 'Withdrawal transaction successful.',
    'data_analysis': {
        'total_bill': f"Total bill amount is",
        'profit':"Deposit interest rate",
        'total_rate_now': 'Total current exchange rate is',
        'total_deposit': 'Total deposit amount',
        'total_withdraw': 'Total withdrawal amount',
        'total_revenue': 'Total revenue',
        'unrefund_money': 'Unrefunded amount',
        'refund_money': 'Amount refunded',
        'total_revenue_USD': 'Total revenue calculated in USD'
    },
    'trade_fail2': "An error occurred during transaction processing.",
    'excel_config': {
        'gd':"Transaction",
        'user':'Implementer',
        'time': "Time",
        'money': 'Amount',
        'money_withdraw': 'Withdrawal amount',
        'total_money': 'Total amount',
        'total_withdraw': 'Total withdrawal',
        'value': 'Value',
        'total_revenue': 'Total revenue',
        'today_analysis': "Today's statistics",
        'fail_excel': "An error occurred while exporting Excel file.",
    },
    'refunds_message_success':"Refund successful",
    'refunds_message_fail':"Refund failed, The refund amount must be less than or equal to the amount not yet refunded.",
    'success_change':"Successfully changed!",
    'change_rate': 'Exchange rate has been changed to',
    'value_er': "Even the exchange rate you entered is wrong :)))",
    'update': 'Upgrade successful. The system has been updated,\n Default exchange rate is 25,000 VND \n For more information about this robot, please contact @Zhiping_Yin \n Thank you.',
    'help': 'How to use this robot:\n\t/start: to start re-calculating bill from the beginning of the day\n\t/end: to end a day\n\t/file: to export total deposit amount of the day to a file\n\t/rate {A}: $US 1 = A VND \n\t+{number}: to add deposit bill\n\t-{number}: to add withdrawal bill\n\t/refunds A VND: refunds A VND\n\n\t/help: to view help\n\t/update: to update the system\n\t/set {language}: (Supported languages: Vietnamese: vi, English: en, Chinese: zh)\n\t/dir A: Change the deposit interest rate to A%.',
    'end': 'The bot has stopped. Please use the command /start to restart.',
    'syntax_er': 'The syntax used does not exist, please use /help for more information.',
    'erorr_undefined': 'Undefined error',
    'change_language': 'Language changed successfully.',
    'send_money':'Send money',
    'withdraw_money':'Withdraw Money',
    'refunds_money':'Refunds',
    'gd':"Transaction",
    'error_float':'Use dots to separate decimal numbers.'
}

lang_cn = {
    "flag":f'{china_flag}',
    'money': 'VND',
    'hello': f"你好 {china_flag}！我是B02（尹志平）制作的机器人，用于记录一天中的存取交易。\n感谢您的使用。很高兴为您服务！",
    'trade_fail': "语法错误",
    'trade_success_pl': '存款成功。',
    'trade_success_sub': '提款成功。',
    'data_analysis': {
        'total_bill': f"账单总额为",
        'profit':"存款利率",
        'total_rate_now': '当前总汇率为',
        'total_deposit': '总存款金额',
        'total_withdraw': '总提款金额',
        'total_revenue': '总收入',
        'unrefund_money': '未退款金额',
        'refund_money': '退款金额',
        'total_revenue_USD': '总收入（以美元计算）'
    },
    'trade_fail2': "交易处理过程中发生了错误。",
    'excel_config': {
        'gd':"交易",
        'user':'实施者',
        'time': "时间",
        'money': '金额',
        'money_withdraw': '提款金额',
        'total_money': '总金额',
        'total_withdraw': '总提款',
        'value': '值',
        'total_revenue': '总收入',
        'today_analysis': '今天的统计',
        'fail_excel': "导出Excel文件时发生错误。",
    },
    'refunds_message_success':"成功退款",
    'refunds_message_fail':"成功退款, 退款金额必须小于或等于尚未退款的金额。",
    'success_change':"修改成功！",
    'change_rate': '汇率已更改为',
    'value_er': '输入的汇率也是错的 :))) NGU!',
    'update': '已成功升级。 系统已全部更新，\n 默认汇率为25.000 VND \n 如需了解更多有关此机器人的信息，请联系@Zhiping_Yin \n 谢谢。',
    'help': '如何使用本机器人：\n\t/start: 开始重新计算当天的账单\n\t/end: 结束一天\n\t/file: 将当天的总存款金额导出到文件\n\t/rate {A}: $US 1 = A VND \n\t+{数}: 添加存款账单\n\t-{数}: 添加提款账单\n\t/refunds A VND: 退款 A VND\n\n\t/help: 查看帮助\n\t/update: 更新系统\n\t/set {语言}: (支持的语言：越南语: vi、英语: en、中文: zh)\n\t/dir A: 将存款利率更改为A%。',
    'end': '机器人已停止。 请使用命令/start重新启动。',
    'syntax_er': '使用的语法不存在，请使用/help获取更多信息。',
    'erorr_undefined': '未定义错误',
    'change_language': '语言成功更改',
    'send_money':'发送钱',
    'withdraw_money':'提取现金',
    'refunds_money':'退款',
    'gd':"交易",
    'error_float':'使用点来分隔小数。'
    
}
# Tỷ giá tiền
exchange_rate_default = 25000
current_language = lang_vi
def format_currency(amount):
    # Xử lý số thực
    if isinstance(amount, float):
        formatted_amount = "{:,.2f}".format(amount).rstrip("0").rstrip(".")
        return formatted_amount

    # Xử lý số nguyên
    if isinstance(amount, int):
        return "{:,.0f}".format(amount)

    return None
def format_currency_vnd(amount):
    formatted_amount = "{:,.0f}".format(round(amount))
    return formatted_amount
# Hàm reset dữ liệu giao dịch
def reset_transactions(message):
    global transactions
    try:
        transactions[message.chat.id]['transactions'] = []
        transactions[message.chat.id]['refund_money'] = 0 
        transactions[message.chat.id]['total_revenue_'] = 0 
    except:
        pass
# Hàm xử lý giao dịch
def handle_transaction(message, current_language, name_user):
    global num_plus_trade, num_sub_trade, total_revenue_, exchange_rate_default
    rate = transactions[message.chat.id]['exchange_rate_default']
    rate = float(rate)
    try:
        text = message.text.strip()
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        # Xử lý số tiền
        currency = current_language['money']
        transactions_list = []
        try:
                data = transactions[message.chat.id]
                try:
                    transactions_list = transactions[message.chat.id]['transactions']
                except:
                    transactions_list = []
                    transactions[message.chat.id]['transactions'] = []
        except:
            transactions_list = []
            try:
                transactions_list = transactions[message.chat.id]['transactions']
            except:
                transactions[message.chat.id]['transactions'] = transactions_list
        if text.startswith('+') or text.startswith('-') or text.startswith('/refunds'):
            amount = int(text.replace('+', '').replace('-', '').replace('/refunds', ''))  # Loại bỏ dấu '+' hoặc '-' và chuyển sang số thực

            
            if amount == 0:  
                bot.reply_to(message, f"{current_time} {format_currency(amount)} {currency} - {current_language['trade_fail']}")
            else:
                if text.startswith('+'):
                    transactions_list.append(("plus",current_time, amount, currency, name_user))
                    bot.reply_to(message, f"{current_time} {format_currency(amount)} {currency} - {current_language['trade_success_pl']}")

                elif text.startswith('-'):
                    transactions_list.append(("sub",current_time, amount, currency,name_user))
                    bot.reply_to(message, f"{current_time} {format_currency(amount)} {currency} - {current_language['trade_success_sub']}")
                elif text.startswith('/refunds'):
                    active_total_revenue_ = transactions[message.chat.id]['total_revenue_']
                    if int(active_total_revenue_) >= int(amount):
                        transactions_list.append(("refunds",current_time, amount, currency,name_user))
                        bot.reply_to(message, f"{current_time} {format_currency(amount)} {currency} - {current_language['refunds_message_success']}")
                    else:
                        # transactions_list.append(("refunds",current_time, amount, currency,name_user))
                        return bot.reply_to(message, f"{current_time} {format_currency(amount)} {currency} - {current_language['refunds_message_fail']}")

        total_deposit = sum([item[2] for item in transactions_list if item[0] == "plus"])
        total_withdraw = sum([item[2] for item in transactions_list if item[0] == 'sub'])
        total_refunds = sum([item[2] for item in transactions_list if item[0] == 'refunds'])

        active_profit = transactions[message.chat.id]['profit']
        total_deposit_ = total_deposit - (total_deposit * active_profit / 100)
        total_withdraw_ = total_withdraw
        total_deposit_usd = format_currency(total_deposit_ / rate)
        total_withdraw_usd = format_currency(total_withdraw_ / rate)

        transactions[message.chat.id]['total_revenue_'] = total_deposit_ - total_withdraw_
        active_total_revenue_ = transactions[message.chat.id]['total_revenue_']
        total_revenue_usd = active_total_revenue_ / rate

        #tiền chưa hoàn trả bằng tổng tiền (total_revenue) trừ tiền đã hoàn trả
        unrefund_money = active_total_revenue_ - total_refunds
        unrefund_money_usd =unrefund_money / rate

        
        str_plus_trade = ''
        str_sub_trade = ''
        num_plus_trade = 0
        num_sub_trade = 0
        for transaction in transactions_list:
            if transaction[0] == 'plus':
                str_plus_trade += f"   {transaction[1]}  {blue_color_start}+{format_currency_vnd(transaction[2])}{blue_color_end} {transaction[3]}\n"
                num_plus_trade += 1
            elif transaction[0] =='sub':
                str_sub_trade += f"   {transaction[1]}  {blue_color_start}-{format_currency_vnd(transaction[2])}{blue_color_end} {transaction[3]}\n"
                num_sub_trade += 1
        
        context = f"{current_language['send_money']} ({num_plus_trade} {current_language['gd']})\n{str_plus_trade}\n{current_language['withdraw_money']} ({num_sub_trade} {current_language['gd']})\n{str_sub_trade}"
# <u>This is underlined text</u>
        data = current_language['data_analysis']
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton('Liên hệ Admin', url='https://t.me/AI_technology_production'))

        transactions[message.chat.id]['transactions'] = transactions_list
        bot.reply_to(message, f"<u>{data['total_bill']} ({num_plus_trade+num_sub_trade})</u>\n\n"
                               f"{context}\n"
                               f"➤ {data['total_deposit']}: {blue_color_start}{format_currency_vnd(total_deposit)}{blue_color_end} {currency} | {blue_color_start}{format_currency(total_deposit/rate)}{blue_color_end} USDT\n"
                               f"➤{data['total_withdraw']}: {blue_color_start}{format_currency_vnd(total_withdraw)}{blue_color_end} {currency} | {blue_color_start}{format_currency(total_withdraw/rate)}{blue_color_end} USDT\n\n"
                               f"{data['profit']}: <b>{active_profit}</b> %\n"
                               f"{data['total_rate_now']}: <b>{format_currency(rate)}</b> {currency} = <b>1</b> USDT\n\n"
                               f"➤ {data['total_deposit']}: {blue_color_start}{format_currency_vnd(total_deposit_)}{blue_color_end} {currency} | {blue_color_start}{total_deposit_usd}{blue_color_end} USDT\n"
                               f"➤{data['total_withdraw']}: {blue_color_start}{format_currency_vnd(total_withdraw_)}{blue_color_end} {currency} | {blue_color_start}{total_withdraw_usd}{blue_color_end} USDT\n\n"
                               f"{data['total_revenue']}:  {blue_color_start}{format_currency_vnd(active_total_revenue_)}{blue_color_end} {currency} | {blue_color_start}{format_currency(total_revenue_usd)}{blue_color_end} USDT\n"
                               f"{data['refund_money']}: {blue_color_start}{format_currency_vnd(total_refunds)}{blue_color_end} {currency} | {blue_color_start}{format_currency(total_refunds / rate)}{blue_color_end} USDT\n"
                               f"{data['unrefund_money']}: {blue_color_start}{format_currency_vnd(unrefund_money)}{blue_color_end} {currency} | {blue_color_start}{format_currency(unrefund_money_usd)}{blue_color_end} USDT\n"
                               f"{current_language['flag']} <u>Created by B02</u>", parse_mode='HTML', reply_markup=keyboard)

    except Exception as e:
        print(e)
        bot.reply_to(message, current_language['trade_fail2'])



# Hàm xuất file Excel
def export_excel(message, current_language):
    try:
        data = current_language['excel_config']

        wb = openpyxl.Workbook()
        sheet = wb.active

        gray_fill = openpyxl.styles.PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        white_font = openpyxl.styles.Font(color="FFFFFF")
        for cell in ["A1", "B1", "C1", "D1", "E1"]:
            sheet[cell].fill = gray_fill
            sheet[cell].font = white_font

        # Đặt tiêu đề cho các cột
        sheet['A1'] = data['time']
        sheet['B1'] = current_language['send_money']
        sheet['C1'] = current_language['withdraw_money']
        sheet['D1'] = current_language['refunds_money']
        sheet['E1'] = data['user']

        try:
            transactions_list = transactions[message.chat.id]['transactions']
        except:
            return bot.reply_to(message, f'not information!')
        column_widths = [15, 30, 30, 30, 25]  
        for i, width in enumerate(column_widths):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width
        for i, (status, time, transaction, curency, person) in enumerate(transactions_list, start=2):
            if status == 'plus':
                sheet[f'A{i}'] = time
                sheet[f'B{i}'] = f"+{format_currency_vnd(transaction)} {curency}"
                sheet[f'E{i}'] = person
            elif status == 'sub':
                sheet[f'A{i}'] = time
                sheet[f'C{i}'] = f"-{format_currency_vnd(transaction)} {curency}"
                sheet[f'E{i}'] = person
            elif status == 'refunds':
                sheet[f'A{i}'] = time
                sheet[f'D{i}'] = f"{format_currency_vnd(transaction)} {curency}"
                sheet[f'E{i}'] = person

        # Lưu file Excel tạm thời
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            wb.save(temp_file.name)

            # Gửi file Excel tạm thời qua bot
            with open(temp_file.name, 'rb') as file:
                bot.send_document(message.chat.id, file, caption=data['today_analysis'])

    except Exception as e:
        print(e)
        bot.reply_to(message, data['fail_excel'])

# Hàm thay đổi tỷ giá
def change_exchange_rate_default(message, new_rate, current_language):
    global exchange_rate_default
    transactions[message.chat.id]['exchange_rate_default'] = new_rate
    bot.reply_to(message, f"{current_language['change_rate']} {new_rate} VND.")

# Lệnh nhập giao dịch hoặc yêu cầu xuất file
@bot.message_handler(func=lambda message: True)
def handle_message(message):
    user = message.from_user
    name_user = ''
    name_user += f"{user.first_name}"
    name_user += f" {user.last_name}" if user.last_name else ""
    global current_language, refund_money, refund_money, exchange_rate_from_USD_to_VND, exchange_rate_default, profit


    active_language = ''
    active_profit = ''
    try:
        data_from_chat_id = transactions[message.chat.id]
        try:
            active_language = data_from_chat_id['language']
            active_profit = data_from_chat_id['profit']
            active_refund_money = data_from_chat_id['refund_money']
            active_total_revenue_ = data_from_chat_id['total_revenue_']
            active_exchange_rate_default = data_from_chat_id['exchange_rate_default']
        except:
            data_from_chat_id['language'] = current_language
            data_from_chat_id['profit'] = profit
            data_from_chat_id['refund_money'] = refund_money
            data_from_chat_id['total_revenue_'] = total_revenue_
            data_from_chat_id['exchange_rate_default'] = exchange_rate_default
    except:
        data = {}
        data['language'] = current_language
        data['profit'] = profit
        data['refund_money'] = refund_money
        data['total_revenue_'] = total_revenue_
        data['exchange_rate_default'] = exchange_rate_default
        transactions[message.chat.id] = data

    active_language = transactions[message.chat.id]['language']
    active_profit = transactions[message.chat.id]['profit']
    active_refund_money = transactions[message.chat.id]['refund_money']
    active_total_revenue_ = transactions[message.chat.id]['total_revenue_']
    active_exchange_rate_default = transactions[message.chat.id]['exchange_rate_default']
    try:
        text = message.text.strip().lower()

        if text == "/start":
            bot.reply_to(message, active_language['hello'])
            reset_transactions(message)
        elif text == "/file":
            export_excel(message, active_language)
        elif text.startswith(('+', '-')):
            handle_transaction(message, active_language, name_user)
        elif text.startswith("/rate"):
            try:
                new_rate = text.split()[1]
                if int(new_rate) < 1:
                    return bot.reply_to(message, active_language['value_er'])
                change_exchange_rate_default(message, new_rate, active_language)
                handle_transaction(message, active_language, name_user)
            except ValueError:
                bot.reply_to(message, active_language['value_er'])
        elif text == "/update":
            pass
        elif text == "/help":
            bot.reply_to(message, active_language['help'])
        elif text == "/end":
            bot.reply_to(message, active_language['end'])
            reset_transactions(message)
        elif text.startswith("/set"):
            language = text.split(' ')[1].lower()
            if language == 'vi':
                active_language = lang_vi
                transactions[message.chat.id]['language'] = active_language
            elif language == 'en':
                active_language = lang_en
                transactions[message.chat.id]['language'] = active_language
            elif language == 'zh':
                active_language = lang_cn
                transactions[message.chat.id]['language'] = active_language
            else:
                bot.reply_to(message, active_language['syntax_er'])
            bot.reply_to(message, active_language['change_language'])
            bot.reply_to(message, active_language['hello'])
        elif text.startswith("/refunds"):
            
            # refund_money = float(text.split()[1])
            return handle_transaction(message, active_language, name_user)
            
        elif text.startswith("/dir"):
            if "%" in text:
                return bot.reply_to(message, 'not %')
            if "," in text:
                return bot.reply_to(message, active_language['error_float'])
            if float(text.split()[1]) >= 0:
                transactions[message.chat.id]['profit'] = float(text.split()[1])
                bot.reply_to(message, active_language['success_change'])
            return handle_transaction(message, active_language, name_user)
            
        else:
            print(total_revenue_)
        
    except Exception as e:
        print(e)
        bot.reply_to(message, active_language['erorr_undefined'])

# # Khởi động bot
# while True:
#     try:
#         bot.polling()
#     except Exception:
#         pass



bot.infinity_polling(timeout=10, long_polling_timeout = 5)